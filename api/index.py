import os
import sys
import io
import json
import re
import numpy as np
import pandas as pd
import plotly
import plotly.graph_objects as go
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, request, render_template_string, send_file, session, redirect, url_for

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from forecasting import (
    moving_average_forecast,
    weighted_moving_average_forecast,
    exponential_smoothing_forecast,
    holt_winters_forecast,
    holt_winters_components,
    linear_trend_forecast,
    seasonal_naive_forecast,
    seasonal_decomposition_forecast,
    seasonal_decomposition_components,
    ensemble_forecast,
    METHOD_DESCRIPTIONS,
)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "farmley-forecast-2024")
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024

FORECAST_METHODS = {
    "Seasonal Naive": seasonal_naive_forecast,
    "Moving Average (3M)": lambda s: moving_average_forecast(s, window=3),
    "Moving Average (6M)": lambda s: moving_average_forecast(s, window=6),
    "Weighted Moving Average": weighted_moving_average_forecast,
    "Exponential Smoothing": exponential_smoothing_forecast,
    "Holt-Winters (Additive)": lambda s: holt_winters_forecast(s, seasonal="add"),
    "Holt-Winters (Multiplicative)": lambda s: holt_winters_forecast(s, seasonal="mul"),
    "Linear Trend + Seasonality": linear_trend_forecast,
    "Seasonal Decomposition": seasonal_decomposition_forecast,
    "Ensemble (Best-of-3)": ensemble_forecast,
}

MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

_cached_data = {}


def parse_month_col(col_name):
    m = re.match(r"([A-Za-z]{3})-(\d{2,4})", str(col_name))
    if not m:
        return None
    abbr, yr = m.group(1), int(m.group(2))
    if yr < 100:
        yr += 2000
    try:
        mi = MONTH_ABBR.index(abbr.capitalize())
    except ValueError:
        return None
    return (yr, mi)


def compute_forecast_months(month_cols, n=9):
    parsed = parse_month_col(month_cols[-1])
    if not parsed:
        yr, mi = 2026, 3
    else:
        yr, mi = parsed
    result = []
    for _ in range(n):
        mi += 1
        if mi > 11:
            mi = 0
            yr += 1
        result.append(f"{MONTH_ABBR[mi]}-{yr}")
    return result


def backtest_method(hist_values, method_fn, n_test=3):
    if len(hist_values) < n_test + 3:
        return None, None
    train = hist_values[:-n_test]
    test = hist_values[-n_test:]
    train_s = pd.Series(train, index=pd.date_range("2020-01-01", periods=len(train), freq="MS"))
    try:
        pred = np.array([float(v) for v in method_fn(train_s)[:n_test]])
        mae = float(np.mean(np.abs(pred - test)))
        safe = np.where(test == 0, 1, test)
        mape = float(np.mean(np.abs((pred - test) / safe))) * 100
        return round(mae, 2), round(mape, 1)
    except Exception:
        return None, None


def build_chart_json(item_name, metric, hist_values, forecasts_dict, month_cols, forecast_months):
    fig = go.Figure()
    x_actual = list(month_cols)
    fig.add_trace(go.Scatter(
        x=x_actual, y=[float(v) for v in hist_values],
        mode="lines+markers", name="Actual",
        line=dict(color="#6366f1", width=3), marker=dict(size=7, symbol="circle"),
        fill="tozeroy", fillcolor="rgba(99,102,241,0.06)",
    ))
    colors = ["#f59e0b", "#10b981", "#ef4444", "#8b5cf6", "#ec4899",
              "#06b6d4", "#84cc16", "#f97316", "#14b8a6", "#a855f7"]
    all_x = x_actual + forecast_months
    for i, (method, vals) in enumerate(forecasts_dict.items()):
        all_y = [float(v) for v in hist_values] + [float(v) for v in vals]
        fig.add_trace(go.Scatter(
            x=all_x, y=all_y, mode="lines+markers", name=method,
            line=dict(color=colors[i % len(colors)], width=2, dash="dash"),
            marker=dict(size=5),
        ))
    last_idx = len(x_actual) - 1
    fig.add_shape(type="line", x0=last_idx, x1=last_idx, y0=0, y1=1, yref="paper",
                  line=dict(color="#cbd5e1", width=2, dash="dot"))
    fig.add_annotation(x=last_idx, y=1.05, yref="paper", text="Actual | Forecast",
                       showarrow=False, font=dict(size=10, color="#94a3b8", family="Inter"))
    fig.update_layout(
        xaxis_title="Month", yaxis_title=metric,
        hovermode="x unified", height=440, template="plotly_white",
        font=dict(family="Inter, system-ui, sans-serif", size=12, color="#334155"),
        legend=dict(orientation="h", yanchor="bottom", y=-0.3, font=dict(size=11)),
        margin=dict(b=90, t=20, l=60, r=20),
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(gridcolor="#f1f5f9", linecolor="#e2e8f0"),
        yaxis=dict(gridcolor="#f1f5f9", linecolor="#e2e8f0"),
    )
    return json.dumps(fig, cls=plotly.utils.PlotlyJSONEncoder)


def create_excel_output(df, month_cols, forecast_months, selected_methods):
    wb = openpyxl.Workbook()
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    ffill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    afill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    valfill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    bold = Font(bold=True)

    n_hist = len(month_cols)
    fc_start_col = 4 + n_hist
    first_data_col = 4

    # Ensemble needs these 3 sheets to exist for cross-sheet MEDIAN references
    ensemble_deps = ["Seasonal Naive", "Holt-Winters (Multiplicative)", "Linear Trend + Seasonality"]
    if "Ensemble (Best-of-3)" in selected_methods:
        for dep in ensemble_deps:
            if dep not in selected_methods:
                selected_methods.append(dep)

    for method_name in selected_methods:
        method_fn = FORECAST_METHODS[method_name]
        ws = wb.create_sheet(title=method_name[:31])

        headers = ["Item_Name", "Group", "Metric"] + list(month_cols) + forecast_months + ["MAE", "MAPE (%)"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            ws.cell(row=r_idx, column=1, value=row["Item_Name"]).border = border
            ws.cell(row=r_idx, column=2, value=row.get("New MIS ITEM Group", "")).border = border
            ws.cell(row=r_idx, column=3, value=row["Metric"]).border = border

            hist = []
            for ci, m in enumerate(month_cols, first_data_col):
                v = float(row[m])
                cell = ws.cell(row=r_idx, column=ci, value=v)
                cell.fill = afill
                cell.border = border
                cell.number_format = '#,##0.00'
                hist.append(v)

            s = pd.Series(hist, index=pd.date_range("2020-01-01", periods=len(hist), freq="MS"))
            try:
                fc_values = [float(v) for v in method_fn(s)[:9]]
            except Exception:
                fc_values = [0.0] * 9

            _write_forecast_formulas(ws, r_idx, method_name, n_hist, first_data_col,
                                     fc_start_col, fc_values, hist, ffill, border)

            mae_col = fc_start_col + len(forecast_months)
            mape_col = mae_col + 1
            mae, mape = backtest_method(np.array(hist), method_fn)
            if mae is not None:
                cell = ws.cell(row=r_idx, column=mae_col, value=mae)
                cell.border = border
                cell.number_format = '#,##0.00'
                cell.fill = valfill
            if mape is not None:
                cell = ws.cell(row=r_idx, column=mape_col, value=mape)
                cell.border = border
                cell.number_format = '#,##0.0'
                cell.fill = valfill

        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 18
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        legend_row = len(df) + 3
        ws.cell(row=legend_row, column=1, value="FORMULA LEGEND:").font = Font(bold=True, size=11, color="1F4E79")
        ws.cell(row=legend_row + 1, column=1, value="Blue cells").fill = afill
        ws.cell(row=legend_row + 1, column=2, value="= Actual historical data (input)")
        ws.cell(row=legend_row + 2, column=1, value="Yellow cells").fill = ffill
        ws.cell(row=legend_row + 2, column=2, value="= Forecasted values (computed by formulas)")
        ws.cell(row=legend_row + 3, column=1, value="Green cells").fill = valfill
        ws.cell(row=legend_row + 3, column=2, value="= Accuracy metrics (MAE / MAPE)")
        ws.cell(row=legend_row + 5, column=1, value="METHOD:").font = bold
        ws.cell(row=legend_row + 5, column=2, value=method_name).font = bold
        formula_explanation = _get_method_formula_text(method_name, n_hist, first_data_col)
        ws.cell(row=legend_row + 6, column=1, value="Formula logic:")
        ws.cell(row=legend_row + 6, column=2, value=formula_explanation)
        ws.row_dimensions[legend_row + 6].height = 80

    expl = wb.create_sheet(title="Method Explanations")
    expl.column_dimensions["A"].width = 30
    expl.column_dimensions["B"].width = 100
    expl.cell(row=1, column=1, value="Method").font = Font(bold=True, size=12)
    expl.cell(row=1, column=2, value="Explanation & Formula").font = Font(bold=True, size=12)
    for i, (mn, md) in enumerate(METHOD_DESCRIPTIONS.items(), 2):
        expl.cell(row=i, column=1, value=mn).font = Font(bold=True)
        expl.cell(row=i, column=2, value=md.strip())
        expl.row_dimensions[i].height = 80

    summ = wb.create_sheet("Forecast Summary", 0)
    summ_headers = ["Item_Name", "Group", "Metric"] + forecast_months
    for c, h in enumerate(summ_headers, 1):
        cell = summ.cell(row=1, column=c, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    summ_row = 2
    for method_name in selected_methods:
        method_fn = FORECAST_METHODS[method_name]
        summ.cell(row=summ_row, column=1, value=f"-- {method_name} --").font = Font(bold=True, size=11, color="1F4E79")
        summ.merge_cells(start_row=summ_row, start_column=1, end_row=summ_row, end_column=len(summ_headers))
        summ_row += 1
        for _, row in df.iterrows():
            hist = [float(row[m]) for m in month_cols]
            s = pd.Series(hist, index=pd.date_range("2020-01-01", periods=len(hist), freq="MS"))
            try:
                fc = [float(v) for v in method_fn(s)[:9]]
            except Exception:
                fc = [0.0] * 9
            summ.cell(row=summ_row, column=1, value=row["Item_Name"]).border = border
            summ.cell(row=summ_row, column=2, value=row.get("New MIS ITEM Group", "")).border = border
            summ.cell(row=summ_row, column=3, value=row["Metric"]).border = border
            for fi, fv in enumerate(fc):
                cell = summ.cell(row=summ_row, column=4 + fi, value=round(fv, 2))
                cell.fill = ffill
                cell.border = border
                cell.number_format = '#,##0.00'
            summ_row += 1
        summ_row += 1

    for c in range(1, len(summ_headers) + 1):
        summ.column_dimensions[get_column_letter(c)].width = 18
    summ.auto_filter.ref = f"A1:{get_column_letter(len(summ_headers))}1"

    # ── BEST METHOD (OPTIMAL) TAB ──
    best_ws = wb.create_sheet(title="Best Method (Optimal)", index=0)
    bestfill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    bestfill_hdr = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    best_hfont = Font(color="FFFFFF", bold=True, size=11)
    goldfill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")

    best_headers = ["Item_Name", "Group", "Metric", "Best Method", "MAE", "MAPE (%)"] + forecast_months
    for c, h in enumerate(best_headers, 1):
        cell = best_ws.cell(row=1, column=c, value=h)
        cell.fill = bestfill_hdr
        cell.font = best_hfont
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    all_method_names = list(FORECAST_METHODS.keys())
    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        hist = [float(row[m]) for m in month_cols]
        s = pd.Series(hist, index=pd.date_range("2020-01-01", periods=len(hist), freq="MS"))

        best_method = None
        best_mape = float("inf")
        best_mae = None
        for mname in all_method_names:
            fn = FORECAST_METHODS[mname]
            mae, mape = backtest_method(np.array(hist), fn)
            if mape is not None and mape < best_mape:
                best_mape = mape
                best_mae = mae
                best_method = mname

        if best_method is None:
            best_method = "Ensemble (Best-of-3)"
            best_mae, best_mape = None, None

        best_ws.cell(row=r_idx, column=1, value=row["Item_Name"]).border = border
        best_ws.cell(row=r_idx, column=2, value=row.get("New MIS ITEM Group", "")).border = border
        best_ws.cell(row=r_idx, column=3, value=row["Metric"]).border = border
        cell = best_ws.cell(row=r_idx, column=4, value=best_method)
        cell.border = border
        cell.fill = goldfill
        cell.font = Font(bold=True, color="5D4037")
        if best_mae is not None:
            cell = best_ws.cell(row=r_idx, column=5, value=best_mae)
            cell.border = border
            cell.number_format = '#,##0.00'
            cell.fill = bestfill
        if best_mape is not None:
            cell = best_ws.cell(row=r_idx, column=6, value=best_mape)
            cell.border = border
            cell.number_format = '#,##0.0'
            cell.fill = bestfill

        # Write forecast formulas referencing the best method's sheet
        best_sheet_title = best_method[:31]
        if best_sheet_title in wb.sheetnames:
            for fi in range(9):
                col = 7 + fi
                src_col_letter = get_column_letter(fc_start_col + fi)
                cell = best_ws.cell(row=r_idx, column=col)
                cell.value = f"='{best_sheet_title}'!{src_col_letter}{r_idx}"
                cell.fill = goldfill
                cell.border = border
                cell.number_format = '#,##0.00'
        else:
            # Method sheet doesn't exist — write static values
            fn = FORECAST_METHODS[best_method]
            try:
                fc = [float(v) for v in fn(s)[:9]]
            except Exception:
                fc = [0.0] * 9
            for fi in range(9):
                cell = best_ws.cell(row=r_idx, column=7 + fi, value=round(fc[fi], 2))
                cell.fill = goldfill
                cell.border = border
                cell.number_format = '#,##0.00'

    for c in range(1, len(best_headers) + 1):
        best_ws.column_dimensions[get_column_letter(c)].width = 18
    best_ws.auto_filter.ref = f"A1:{get_column_letter(len(best_headers))}1"

    legend_r = len(df) + 3
    best_ws.cell(row=legend_r, column=1, value="BEST METHOD TAB").font = Font(bold=True, size=12, color="1B5E20")
    best_ws.cell(row=legend_r + 1, column=1, value="This tab automatically selects the forecasting method with the lowest MAPE for each item.")
    best_ws.cell(row=legend_r + 2, column=1, value="Forecast cells contain formulas referencing the winning method's sheet.")
    best_ws.cell(row=legend_r + 3, column=1, value="All 10 methods are evaluated per item; the one with lowest MAPE% wins.")

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _write_forecast_formulas(ws, r, method_name, n_hist, data_col, fc_col, fc_values, hist, ffill, border):
    last_data_letter = get_column_letter(data_col + n_hist - 1)
    first_data_letter = get_column_letter(data_col)

    if "Moving Average (3M)" in method_name:
        for fi in range(9):
            col = fc_col + fi
            cell = ws.cell(row=r, column=col)
            if fi == 0:
                c1 = get_column_letter(data_col + n_hist - 3)
                c2 = get_column_letter(data_col + n_hist - 2)
                c3 = get_column_letter(data_col + n_hist - 1)
                cell.value = f"=AVERAGE({c1}{r},{c2}{r},{c3}{r})"
            else:
                c1 = get_column_letter(col - 3)
                c2 = get_column_letter(col - 2)
                c3 = get_column_letter(col - 1)
                cell.value = f"=AVERAGE({c1}{r},{c2}{r},{c3}{r})"
            cell.fill = ffill
            cell.border = border
            cell.number_format = '#,##0.00'

    elif "Moving Average (6M)" in method_name:
        for fi in range(9):
            col = fc_col + fi
            cell = ws.cell(row=r, column=col)
            if fi == 0:
                c_start = get_column_letter(data_col + n_hist - 6)
                c_end = get_column_letter(data_col + n_hist - 1)
                cell.value = f"=AVERAGE({c_start}{r}:{c_end}{r})"
            else:
                c_start = get_column_letter(col - 6)
                c_end = get_column_letter(col - 1)
                cell.value = f"=AVERAGE({c_start}{r}:{c_end}{r})"
            cell.fill = ffill
            cell.border = border
            cell.number_format = '#,##0.00'

    elif "Weighted Moving Average" in method_name:
        for fi in range(9):
            col = fc_col + fi
            cell = ws.cell(row=r, column=col)
            if fi == 0:
                c1 = get_column_letter(data_col + n_hist - 5)
                c2 = get_column_letter(data_col + n_hist - 4)
                c3 = get_column_letter(data_col + n_hist - 3)
                c4 = get_column_letter(data_col + n_hist - 2)
                c5 = get_column_letter(data_col + n_hist - 1)
            else:
                c1 = get_column_letter(col - 5)
                c2 = get_column_letter(col - 4)
                c3 = get_column_letter(col - 3)
                c4 = get_column_letter(col - 2)
                c5 = get_column_letter(col - 1)
            cell.value = f"={c1}{r}*0.10+{c2}{r}*0.15+{c3}{r}*0.20+{c4}{r}*0.25+{c5}{r}*0.30"
            cell.fill = ffill
            cell.border = border
            cell.number_format = '#,##0.00'

    elif "Seasonal Naive" in method_name:
        for fi in range(9):
            col = fc_col + fi
            cell = ws.cell(row=r, column=col)
            src_col = data_col + (fi % n_hist)
            src_letter = get_column_letter(src_col)
            cell.value = f"={src_letter}{r}"
            cell.fill = ffill
            cell.border = border
            cell.number_format = '#,##0.00'

    elif "Exponential Smoothing" in method_name:
        alpha = _find_best_alpha(hist)
        alpha_col = fc_col + 10
        ws.cell(row=r, column=alpha_col, value=round(alpha, 3))
        ws.cell(row=1, column=alpha_col, value="Alpha").font = Font(bold=True, size=9, color="888888")
        alpha_letter = get_column_letter(alpha_col)
        last_actual = get_column_letter(data_col + n_hist - 1)
        avg_letter_start = get_column_letter(data_col)
        avg_letter_end = get_column_letter(data_col + n_hist - 1)
        for fi in range(9):
            col = fc_col + fi
            cell = ws.cell(row=r, column=col)
            cell.value = f"={alpha_letter}{r}*{last_actual}{r}+(1-{alpha_letter}{r})*AVERAGE({avg_letter_start}{r}:{avg_letter_end}{r})"
            cell.fill = ffill
            cell.border = border
            cell.number_format = '#,##0.00'

    elif "Linear Trend" in method_name:
        si_start_col = fc_col + 11
        indices = _compute_seasonal_indices(hist, n_hist)
        for si_i, si_v in enumerate(indices):
            ws.cell(row=r, column=si_start_col + si_i, value=round(si_v, 4))
        if r == 2:
            for si_i in range(n_hist):
                ws.cell(row=1, column=si_start_col + si_i, value=f"SI_{si_i+1}").font = Font(size=8, color="888888")

        seq_start = get_column_letter(data_col)
        seq_end = get_column_letter(data_col + n_hist - 1)
        for fi in range(9):
            col = fc_col + fi
            cell = ws.cell(row=r, column=col)
            trend_x = n_hist + fi + 1
            si_col_letter = get_column_letter(si_start_col + ((n_hist + fi) % n_hist))
            cell.value = f'=TREND({seq_start}{r}:{seq_end}{r},ROW(INDIRECT("1:{n_hist}")),{trend_x})*{si_col_letter}{r}'
            cell.fill = ffill
            cell.border = border
            cell.number_format = '#,##0.00'

    elif "Holt-Winters" in method_name:
        seasonal_type = "mul" if "Multiplicativ" in method_name else "add"
        s = pd.Series(hist, index=pd.date_range("2020-01-01", periods=len(hist), freq="MS"))
        level, trend, season, period, stype = holt_winters_components(s, seasonal_type)
        # Helper columns: Level, Trend, Seasonal[0..period-1]
        helper_start = fc_col + 11
        ws.cell(row=r, column=helper_start, value=round(level, 4))
        ws.cell(row=r, column=helper_start + 1, value=round(trend, 4))
        for si_i, sv in enumerate(season):
            ws.cell(row=r, column=helper_start + 2 + si_i, value=round(sv, 4))
        if r == 2:
            ws.cell(row=1, column=helper_start, value="HW_Level").font = Font(size=8, color="888888")
            ws.cell(row=1, column=helper_start + 1, value="HW_Trend").font = Font(size=8, color="888888")
            for si_i in range(period):
                ws.cell(row=1, column=helper_start + 2 + si_i, value=f"HW_S{si_i+1}").font = Font(size=8, color="888888")
        lev_col = get_column_letter(helper_start)
        trn_col = get_column_letter(helper_start + 1)
        for fi in range(9):
            col = fc_col + fi
            cell = ws.cell(row=r, column=col)
            h = fi + 1
            si = (n_hist + fi) % period
            s_col = get_column_letter(helper_start + 2 + si)
            if stype == "mul":
                cell.value = f"=({lev_col}{r}+{h}*{trn_col}{r})*{s_col}{r}"
            else:
                cell.value = f"={lev_col}{r}+{h}*{trn_col}{r}+{s_col}{r}"
            cell.fill = ffill
            cell.border = border
            cell.number_format = '#,##0.00'

    elif "Decomposition" in method_name:
        s = pd.Series(hist, index=pd.date_range("2020-01-01", periods=len(hist), freq="MS"))
        intercept, slope, seasonal, period = seasonal_decomposition_components(s)
        helper_start = fc_col + 11
        ws.cell(row=r, column=helper_start, value=round(intercept, 4))
        ws.cell(row=r, column=helper_start + 1, value=round(slope, 6))
        for si_i, sv in enumerate(seasonal):
            ws.cell(row=r, column=helper_start + 2 + si_i, value=round(sv, 4))
        if r == 2:
            ws.cell(row=1, column=helper_start, value="STL_Intcpt").font = Font(size=8, color="888888")
            ws.cell(row=1, column=helper_start + 1, value="STL_Slope").font = Font(size=8, color="888888")
            for si_i in range(period):
                ws.cell(row=1, column=helper_start + 2 + si_i, value=f"STL_S{si_i+1}").font = Font(size=8, color="888888")
        int_col = get_column_letter(helper_start)
        slp_col = get_column_letter(helper_start + 1)
        for fi in range(9):
            col = fc_col + fi
            cell = ws.cell(row=r, column=col)
            future_t = n_hist + fi
            si = (n_hist + fi) % period
            s_col = get_column_letter(helper_start + 2 + si)
            cell.value = f"={int_col}{r}+{slp_col}{r}*{future_t}+{s_col}{r}"
            cell.fill = ffill
            cell.border = border
            cell.number_format = '#,##0.00'

    elif "Ensemble" in method_name:
        sn_sheet = "Seasonal Naive"[:31]
        hw_sheet = "Holt-Winters (Multiplicativ"[:31]
        lt_sheet = "Linear Trend + Seasonality"[:31]
        for fi in range(9):
            col = fc_col + fi
            cell = ws.cell(row=r, column=col)
            fc_letter = get_column_letter(col)
            cell.value = f"=MEDIAN('{sn_sheet}'!{fc_letter}{r},'{hw_sheet}'!{fc_letter}{r},'{lt_sheet}'!{fc_letter}{r})"
            cell.fill = ffill
            cell.border = border
            cell.number_format = '#,##0.00'

    else:
        for fi in range(9):
            col = fc_col + fi
            cell = ws.cell(row=r, column=col, value=round(fc_values[fi], 2))
            cell.fill = ffill
            cell.border = border
            cell.number_format = '#,##0.00'


def _find_best_alpha(hist):
    n = len(hist)
    vals = np.array(hist, dtype=float)
    best_a, best_sse = 0.3, float("inf")
    for a_int in range(5, 96, 5):
        a = a_int / 100.0
        f = [vals[0]]
        for t in range(1, n):
            f.append(a * vals[t-1] + (1-a) * f[-1])
        sse = float(np.sum((vals[1:] - np.array(f[1:])) ** 2))
        if sse < best_sse:
            best_sse, best_a = sse, a
    return best_a


def _compute_seasonal_indices(hist, period):
    vals = np.array(hist, dtype=float)
    n = len(vals)
    t = np.arange(n, dtype=float)
    t_mean, v_mean = np.mean(t), np.mean(vals)
    denom = np.sum((t - t_mean) ** 2)
    slope = np.sum((t - t_mean) * (vals - v_mean)) / max(denom, 1e-10)
    intercept = v_mean - slope * t_mean
    trend = intercept + slope * t
    indices = np.ones(period)
    if np.all(np.abs(trend) > 1e-10):
        ratios = vals / trend
        for i in range(period):
            month_vals = [ratios[j] for j in range(i, n, period)]
            indices[i] = np.mean(month_vals) if month_vals else 1.0
    return indices.tolist()


def _get_method_formula_text(method_name, n_hist, data_col):
    d = get_column_letter(data_col)
    last = get_column_letter(data_col + n_hist - 1)
    if "Moving Average (3M)" in method_name:
        return "Each forecast = AVERAGE of previous 3 cells. First forecast: =AVERAGE(last 3 actuals). Subsequent: rolling window shifts right."
    elif "Moving Average (6M)" in method_name:
        return "Each forecast = AVERAGE of previous 6 cells. First forecast: =AVERAGE(last 6 actuals). Subsequent: rolling window shifts right."
    elif "Weighted Moving Average" in method_name:
        return "WMA = 0.10*fifth_last + 0.15*fourth_last + 0.20*third_last + 0.25*second_last + 0.30*last. Weights sum to 1.0, most recent gets highest weight."
    elif "Seasonal Naive" in method_name:
        return f"Each forecast month = same month from previous year. Apr forecast = Apr actual ({d}), May forecast = May actual, etc."
    elif "Exponential" in method_name:
        return "F = alpha * last_actual + (1-alpha) * avg(all_actuals). Alpha optimized by minimizing SSE. Stored in Alpha column."
    elif "Linear Trend" in method_name:
        return f"F = TREND(actuals, 1..{n_hist}, future_t) * Seasonal_Index[month]. TREND extrapolates linear fit. SI columns hold multiplicative seasonal indices."
    elif "Holt-Winters" in method_name:
        op = "*" if "Multiplicativ" in method_name else "+"
        return f"F = (HW_Level + h*HW_Trend) {op} HW_S[month]. Helper columns store optimized Level, Trend, and Seasonal factors. Click any yellow cell to see the formula."
    elif "Decomposition" in method_name:
        return "F = STL_Intercept + STL_Slope * t + STL_S[month]. Helper columns store trend intercept, slope, and seasonal offsets from centered MA decomposition."
    elif "Ensemble" in method_name:
        return "F = MEDIAN('Seasonal Naive'!cell, 'Holt-Winters (Multiplicativ'!cell, 'Linear Trend + Seasonality'!cell). Cross-sheet MEDIAN formula."
    return f"{method_name}: computed values from Python forecasting engine."


TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Farmley Forecast Dashboard</title>
<script src="https://cdn.plot.ly/plotly-2.35.0.min.js"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>
:root{--primary:#6366f1;--primary-dark:#4f46e5;--primary-light:#818cf8;--primary-bg:rgba(99,102,241,.06);--success:#10b981;--success-bg:#ecfdf5;--warning:#f59e0b;--warning-bg:#fffbeb;--danger:#ef4444;--danger-bg:#fef2f2;--surface:#ffffff;--surface-2:#f8fafc;--border:#e2e8f0;--border-light:#f1f5f9;--text:#0f172a;--text-2:#334155;--text-3:#64748b;--text-4:#94a3b8;--shadow-sm:0 1px 2px rgba(0,0,0,.04);--shadow:0 1px 3px rgba(0,0,0,.06),0 1px 2px rgba(0,0,0,.04);--shadow-md:0 4px 6px -1px rgba(0,0,0,.07),0 2px 4px -2px rgba(0,0,0,.05);--shadow-lg:0 10px 15px -3px rgba(0,0,0,.08),0 4px 6px -4px rgba(0,0,0,.04);--radius:12px;--radius-lg:16px;--radius-xl:20px}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Inter',system-ui,-apple-system,sans-serif;background:linear-gradient(135deg,#f0f2f5 0%,#e8ecf4 50%,#f0f2f5 100%);color:var(--text);-webkit-font-smoothing:antialiased;min-height:100vh}

/* ── ANIMATIONS ── */
@keyframes fadeUp{from{opacity:0;transform:translateY(12px)}to{opacity:1;transform:translateY(0)}}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.7}}
@keyframes shimmer{0%{background-position:-200% 0}100%{background-position:200% 0}}
.fade-up{animation:fadeUp .4s ease-out both}
.fade-up-1{animation-delay:.05s}.fade-up-2{animation-delay:.1s}.fade-up-3{animation-delay:.15s}.fade-up-4{animation-delay:.2s}

/* ── NAV ── */
.nav{background:linear-gradient(135deg,#0f172a 0%,#1e293b 100%);padding:0 32px;height:64px;display:flex;align-items:center;gap:14px;position:sticky;top:0;z-index:50;box-shadow:0 4px 20px rgba(0,0,0,.2);backdrop-filter:blur(12px)}
.nav-logo{width:38px;height:38px;background:linear-gradient(135deg,var(--primary),#8b5cf6);border-radius:10px;display:flex;align-items:center;justify-content:center;color:#fff;font-weight:900;font-size:1.15rem;box-shadow:0 2px 8px rgba(99,102,241,.4)}
.nav h1{color:#fff;font-size:1.15rem;font-weight:700;letter-spacing:-.02em}
.nav-sub{color:#64748b;font-size:.72rem;margin-left:-6px;font-weight:500}
.nav-right{margin-left:auto;display:flex;align-items:center;gap:10px}
.nav-badge{background:linear-gradient(135deg,rgba(99,102,241,.2),rgba(139,92,246,.2));color:#a5b4fc;padding:5px 14px;border-radius:20px;font-size:.72rem;font-weight:600;border:1px solid rgba(99,102,241,.2)}

.wrap{max-width:1520px;margin:0 auto;padding:24px 28px 60px}

/* ── UPLOAD ── */
.upload-wrap{max-width:580px;margin:100px auto}
.upload-card{background:var(--surface);border-radius:var(--radius-xl);box-shadow:var(--shadow-lg);padding:56px 48px;text-align:center;border:1px solid var(--border)}
.upload-card h2{font-size:1.5rem;font-weight:800;margin-bottom:6px;background:linear-gradient(135deg,var(--text),var(--primary));-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.upload-card .sub{color:var(--text-3);font-size:.9rem;margin-bottom:32px}
.drop{border:2px dashed #cbd5e1;border-radius:var(--radius);padding:44px 24px;background:var(--surface-2);transition:.25s;cursor:pointer;position:relative}
.drop:hover,.drop.over{border-color:var(--primary);background:#eef2ff;transform:scale(1.01)}
.drop input{position:absolute;inset:0;opacity:0;cursor:pointer}
.drop h3{font-size:1rem;color:var(--text-2);margin-bottom:6px;font-weight:600}
.drop p{font-size:.82rem;color:var(--text-4)}
#fileName{display:none;margin-top:12px;font-size:.88rem;color:var(--primary);font-weight:600}
.btn{display:inline-flex;align-items:center;justify-content:center;gap:7px;padding:11px 24px;background:linear-gradient(135deg,var(--primary),var(--primary-dark));color:#fff;border:none;border-radius:10px;cursor:pointer;font-size:.88rem;font-weight:600;font-family:inherit;text-decoration:none;transition:all .2s;box-shadow:0 2px 8px rgba(99,102,241,.25)}
.btn:hover{transform:translateY(-1px);box-shadow:0 4px 16px rgba(99,102,241,.35);background:linear-gradient(135deg,var(--primary-light),var(--primary))}
.btn:active{transform:translateY(0)}
.btn-block{width:100%}
.btn-green{background:linear-gradient(135deg,#10b981,#059669);box-shadow:0 2px 8px rgba(16,185,129,.25)}.btn-green:hover{box-shadow:0 4px 16px rgba(16,185,129,.35)}
.btn-ghost{background:transparent;color:var(--primary);border:1px solid var(--border);box-shadow:none}.btn-ghost:hover{background:var(--primary-bg);border-color:var(--primary);box-shadow:none;transform:none}
.btn-sm{padding:5px 11px;font-size:.72rem;border-radius:6px;box-shadow:none}

/* ── LAYOUT ── */
.dash{display:grid;grid-template-columns:320px 1fr;gap:24px}
@media(max-width:1000px){.dash{grid-template-columns:1fr}}

/* ── SIDEBAR ── */
.side{position:sticky;top:80px;max-height:calc(100vh - 96px);overflow-y:auto;display:flex;flex-direction:column;gap:14px;padding-right:4px}
.side::-webkit-scrollbar{width:4px}.side::-webkit-scrollbar-thumb{background:#cbd5e1;border-radius:4px}
@media(max-width:1000px){.side{position:static;max-height:none}}
.panel{background:var(--surface);border-radius:var(--radius);box-shadow:var(--shadow);padding:18px;border:1px solid var(--border);transition:box-shadow .2s}
.panel:hover{box-shadow:var(--shadow-md)}
.panel-title{font-size:.76rem;font-weight:700;color:var(--text);text-transform:uppercase;letter-spacing:.07em;margin-bottom:14px;display:flex;align-items:center;gap:7px}
.panel-title svg{width:15px;height:15px;color:var(--primary)}

/* ── CHECKBOX FILTER ── */
.fgroup{margin-bottom:16px}
.fgroup-label{font-size:.7rem;font-weight:700;color:var(--text-3);text-transform:uppercase;letter-spacing:.06em;margin-bottom:7px}
.fgroup-bar{display:flex;align-items:center;gap:5px;margin-bottom:7px}
.fgroup-bar input[type=text]{flex:1;padding:7px 12px;border:1px solid var(--border);border-radius:8px;font-size:.8rem;font-family:inherit;outline:none;transition:.2s;background:var(--surface)}
.fgroup-bar input[type=text]:focus{border-color:var(--primary);box-shadow:0 0 0 3px rgba(99,102,241,.1)}
.cbox-list{max-height:180px;overflow-y:auto;border:1px solid var(--border);border-radius:10px;background:var(--surface-2);padding:4px 0}
.cbox-list::-webkit-scrollbar{width:4px}.cbox-list::-webkit-scrollbar-thumb{background:#cbd5e1;border-radius:4px}
.cbox-list label{display:flex;align-items:center;gap:9px;padding:6px 12px;font-size:.8rem;color:var(--text-2);cursor:pointer;transition:.15s;user-select:none;border-radius:6px;margin:1px 4px}
.cbox-list label:hover{background:var(--primary-bg);color:var(--primary)}
.cbox-list label.hidden{display:none}
.cbox-list input[type=checkbox]{width:16px;height:16px;accent-color:var(--primary);cursor:pointer;flex-shrink:0;border-radius:4px}
.cbox-list .cname{white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.sel-info{font-size:.7rem;color:var(--text-4);margin-top:5px;font-weight:500}

.sselect{margin-bottom:16px}
.sselect label{display:block;font-size:.7rem;font-weight:700;color:var(--text-3);text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px}
.sselect select{width:100%;padding:9px 12px;border:1px solid var(--border);border-radius:8px;font-size:.84rem;font-family:inherit;background:var(--surface);outline:none;cursor:pointer;appearance:none;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='10' fill='%2394a3b8'%3E%3Cpath d='M5 7L1 3h8z'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 12px center;transition:.2s}
.sselect select:focus{border-color:var(--primary);box-shadow:0 0 0 3px rgba(99,102,241,.1)}

/* ── STATS ── */
.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:12px;margin-bottom:20px}
.st{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:18px 20px;transition:all .2s;position:relative;overflow:hidden}
.st:hover{box-shadow:var(--shadow-md);transform:translateY(-2px)}
.st::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;border-radius:var(--radius) var(--radius) 0 0}
.st:nth-child(1)::before{background:linear-gradient(90deg,var(--primary),#8b5cf6)}
.st:nth-child(2)::before{background:linear-gradient(90deg,var(--success),#34d399)}
.st:nth-child(3)::before{background:linear-gradient(90deg,var(--warning),#fbbf24)}
.st:nth-child(4)::before{background:linear-gradient(90deg,#06b6d4,#22d3ee)}
.st-label{font-size:.68rem;font-weight:600;color:var(--text-4);text-transform:uppercase;letter-spacing:.06em}
.st-val{font-size:1.25rem;font-weight:800;color:var(--text);margin-top:4px;letter-spacing:-.02em}

/* ── CARDS ── */
.card{background:var(--surface);border-radius:var(--radius-lg);box-shadow:var(--shadow);padding:24px;margin-bottom:20px;border:1px solid var(--border);transition:all .25s}
.card:hover{box-shadow:var(--shadow-md)}
.item-card{border-left:4px solid var(--primary);position:relative}
.item-card::before{content:'';position:absolute;top:0;left:-4px;bottom:0;width:4px;background:linear-gradient(180deg,var(--primary),#8b5cf6);border-radius:var(--radius-lg) 0 0 var(--radius-lg)}
.item-head{display:flex;align-items:center;gap:10px;margin-bottom:16px;flex-wrap:wrap}
.item-head h2{font-size:1.05rem;font-weight:700;color:var(--text);margin:0}
.tag{font-size:.68rem;font-weight:600;background:var(--border-light);color:var(--text-3);padding:4px 11px;border-radius:20px;transition:.15s}
.tag:hover{background:#e2e8f0}
.section-lbl{font-size:.82rem;font-weight:700;color:var(--text);margin:22px 0 10px;display:flex;align-items:center;gap:7px}
.section-lbl svg{width:15px;height:15px;color:var(--primary)}

/* ── TABLES ── */
table{width:100%;border-collapse:separate;border-spacing:0;font-size:.8rem}
thead th{background:linear-gradient(180deg,var(--surface-2),#eef0f4);color:var(--text-3);padding:10px 14px;text-align:left;font-weight:600;font-size:.72rem;text-transform:uppercase;letter-spacing:.05em;border-bottom:2px solid var(--border);position:sticky;top:0;z-index:1}
tbody td{padding:10px 14px;border-bottom:1px solid var(--border-light);color:var(--text-2);transition:background .15s}
tbody tr:hover td{background:var(--primary-bg)}
.fc-val{background:#fffbeb;font-weight:600;color:#92400e;font-variant-numeric:tabular-nums}
.scroll-tbl{overflow-x:auto;max-height:420px;overflow-y:auto;border:1px solid var(--border);border-radius:10px}

/* ── BEST METHOD HIGHLIGHT ── */
.best-row td{background:linear-gradient(90deg,rgba(16,185,129,.08),rgba(16,185,129,.03))!important;font-weight:600}
.best-row td:first-child{position:relative;padding-left:22px}
.best-row td:first-child::before{content:'';position:absolute;left:6px;top:50%;transform:translateY(-50%);width:8px;height:8px;background:var(--success);border-radius:50%;box-shadow:0 0 6px rgba(16,185,129,.4)}
.best-badge{display:inline-flex;align-items:center;gap:4px;background:linear-gradient(135deg,#059669,#10b981);color:#fff;padding:3px 10px;border-radius:20px;font-size:.66rem;font-weight:700;letter-spacing:.02em;box-shadow:0 2px 6px rgba(16,185,129,.3)}
.best-badge svg{width:10px;height:10px}

/* ── BADGES ── */
.badge{display:inline-flex;align-items:center;padding:4px 12px;border-radius:20px;font-size:.7rem;font-weight:600;gap:4px}
.badge-good{background:var(--success-bg);color:#065f46}
.badge-fair{background:var(--warning-bg);color:#92400e}
.badge-poor{background:var(--danger-bg);color:#991b1b}

/* ── METHOD TABS ── */
.tabs{display:flex;gap:4px;margin-bottom:16px;flex-wrap:wrap;padding:4px;background:var(--border-light);border-radius:10px}
.tab{padding:8px 16px;cursor:pointer;font-size:.76rem;font-weight:500;color:var(--text-3);border-radius:8px;transition:.2s}
.tab:hover{color:var(--text-2);background:rgba(255,255,255,.5)}
.tab.active{background:var(--surface);color:var(--primary);font-weight:600;box-shadow:var(--shadow-sm)}
.tab-content{display:none}.tab-content.active{display:block}
.method-desc{line-height:1.75;font-size:.86rem;color:var(--text-2)}
.method-desc strong{color:var(--text)}
.method-desc code{background:var(--primary-bg);color:var(--primary);padding:2px 7px;border-radius:5px;font-size:.78rem;font-weight:500}

.guide{margin-top:24px;padding:20px;background:linear-gradient(135deg,var(--surface-2),#f0f4ff);border-radius:var(--radius);border:1px solid var(--border)}
.guide h3{font-size:.92rem;font-weight:700;color:var(--text);margin-bottom:12px}
.guide-note{margin-top:14px;font-size:.8rem;color:var(--text-3);line-height:1.7}
.guide-note strong{color:var(--text-2)}

.empty{text-align:center;padding:72px 24px;color:var(--text-4)}
.empty p{font-size:.92rem}.empty strong{color:var(--text-3)}

footer{text-align:center;padding:32px 24px;color:var(--text-4);font-size:.74rem;border-top:1px solid var(--border);margin-top:16px}
footer a{color:var(--primary);text-decoration:none;font-weight:500}
footer a:hover{text-decoration:underline}
</style>
</head>
<body>

<div class="nav">
    <div class="nav-logo">F</div>
    <div>
        <h1>Farmley Forecast</h1>
        <span class="nav-sub">Sales Forecasting Dashboard</span>
    </div>
    {% if data_loaded %}
    <div class="nav-right">
        <span class="nav-badge">{{total_items}} Products</span>
        <span class="nav-badge">{{n_months}} Months Data</span>
    </div>
    {% endif %}
</div>

{% if not data_loaded %}
<div class="upload-wrap fade-up">
    <div class="upload-card">
        <h2>Upload Sales Data</h2>
        <p class="sub">Excel file with item-wise monthly sales order data</p>
        <form method="post" enctype="multipart/form-data" action="/upload">
            <div class="drop" id="dropZone">
                <svg width="48" height="48" viewBox="0 0 24 24" fill="none" stroke="#94a3b8" stroke-width="1.5" style="margin-bottom:12px"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
                <h3>Click to browse or drag & drop</h3>
                <p>.xlsx or .xls files supported</p>
                <span id="fileName"></span>
                <input type="file" name="file" accept=".xlsx,.xls" required id="fileInput">
            </div>
            <button type="submit" class="btn btn-block" style="margin-top:22px;padding:14px 24px;font-size:.95rem">Upload & Analyze</button>
        </form>
    </div>
</div>
<script>
var dz=document.getElementById('dropZone'),fi=document.getElementById('fileInput'),fn=document.getElementById('fileName');
if(dz){
    ['dragenter','dragover'].forEach(function(e){dz.addEventListener(e,function(ev){ev.preventDefault();dz.classList.add('over')})});
    ['dragleave','drop'].forEach(function(e){dz.addEventListener(e,function(ev){ev.preventDefault();dz.classList.remove('over')})});
    dz.addEventListener('drop',function(e){if(e.dataTransfer.files.length){fi.files=e.dataTransfer.files;fn.style.display='block';fn.textContent=fi.files[0].name}});
    fi.addEventListener('change',function(){if(fi.files.length){fn.style.display='block';fn.textContent=fi.files[0].name}});
}
</script>

{% else %}

<div class="wrap">
<form method="post" action="/forecast" id="mainForm">
<div class="dash">

<!-- ── SIDEBAR ── -->
<div class="side">
    <div class="panel fade-up">
        <div class="panel-title">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5"><polygon points="22 3 2 3 10 12.46 10 19 14 21 14 12.46 22 3"/></svg>
            Filters
        </div>

        <div class="fgroup">
            <div class="fgroup-label">Product Group</div>
            <div class="fgroup-bar">
                <input type="text" placeholder="Search groups..." onkeyup="filterCB(this,'grp')">
                <button type="button" class="btn btn-sm btn-ghost" onclick="checkAll('grp')">All</button>
                <button type="button" class="btn btn-sm btn-ghost" onclick="uncheckAll('grp')">None</button>
            </div>
            <div class="cbox-list" id="grp">
            {% for g in all_groups %}
                <label><input type="checkbox" name="groups" value="{{g}}" {% if g in selected_groups %}checked{% endif %}><span class="cname">{{g}}</span></label>
            {% endfor %}
            </div>
            <div class="sel-info" id="grp_info"></div>
        </div>

        <div class="fgroup">
            <div class="fgroup-label">Item Name</div>
            <div class="fgroup-bar">
                <input type="text" placeholder="Search items..." onkeyup="filterCB(this,'itm')">
                <button type="button" class="btn btn-sm btn-ghost" onclick="checkAll('itm')">All</button>
                <button type="button" class="btn btn-sm btn-ghost" onclick="uncheckAll('itm')">None</button>
            </div>
            <div class="cbox-list" id="itm">
            {% for it in all_items %}
                <label><input type="checkbox" name="items" value="{{it}}" {% if it in selected_items %}checked{% endif %}><span class="cname">{{it}}</span></label>
            {% endfor %}
            </div>
            <div class="sel-info" id="itm_info"></div>
        </div>

        <div class="sselect">
            <label>Metric</label>
            <select name="metric">
            {% for m in all_metrics %}
                <option value="{{m}}" {% if m == selected_metric %}selected{% endif %}>{{m}}</option>
            {% endfor %}
            </select>
        </div>

        <div class="fgroup">
            <div class="fgroup-label">Forecast Methods</div>
            <div class="fgroup-bar">
                <button type="button" class="btn btn-sm btn-ghost" onclick="checkAll('mth')">All</button>
                <button type="button" class="btn btn-sm btn-ghost" onclick="uncheckAll('mth')">None</button>
            </div>
            <div class="cbox-list" id="mth" style="max-height:220px">
            {% for m in all_methods %}
                <label><input type="checkbox" name="methods" value="{{m}}" {% if m in selected_methods %}checked{% endif %}><span class="cname">{{m}}</span></label>
            {% endfor %}
            </div>
            <div class="sel-info" id="mth_info"></div>
        </div>

        <button type="submit" class="btn btn-block" style="margin-top:6px">Apply Filters</button>
    </div>

    <div class="panel fade-up fade-up-1">
        <div class="panel-title">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
            Export
        </div>
        <p style="font-size:.76rem;color:var(--text-4);margin-bottom:10px">Excel with formulas, Best Method tab, MAE/MAPE metrics, one tab per method.</p>
        <a href="/download?methods={{selected_methods|join(',')}}" class="btn btn-green btn-block">Download Forecast Excel</a>
    </div>

    <div class="panel fade-up fade-up-2">
        <a href="/" class="btn btn-ghost btn-block">Upload New File</a>
    </div>
</div>

<!-- ── MAIN ── -->
<div>
    <div class="stats fade-up">
        <div class="st"><div class="st-label">Total Items</div><div class="st-val">{{total_items}}</div></div>
        <div class="st"><div class="st-label">Data Through</div><div class="st-val">{{last_month}}</div></div>
        <div class="st"><div class="st-label">Forecast Range</div><div class="st-val">{{forecast_months[0]}} - {{forecast_months[-1]}}</div></div>
        <div class="st"><div class="st-label">Months of Data</div><div class="st-val">{{n_months}}</div></div>
    </div>

    {% if overall_chart %}
    <div class="card fade-up fade-up-1" style="border-left:4px solid var(--success);position:relative;overflow:hidden">
        <div style="position:absolute;top:0;left:-4px;bottom:0;width:4px;background:linear-gradient(180deg,var(--success),#34d399);border-radius:var(--radius-lg) 0 0 var(--radius-lg)"></div>
        <div class="item-head">
            <h2>Overall Aggregate</h2>
            <span class="tag" style="background:var(--success-bg);color:#065f46">{{results|length}} items combined</span>
            <span class="tag">{{selected_metric}}</span>
        </div>
        <div id="chart_overall" style="width:100%;height:480px"></div>
        <script>Plotly.newPlot('chart_overall',{{overall_chart|safe}}.data,{{overall_chart|safe}}.layout,{responsive:true})</script>
    </div>
    {% endif %}

    {% if group_charts %}
    {% for gc in group_charts %}
    <div class="card fade-up" style="border-left:4px solid var(--warning);position:relative;overflow:hidden">
        <div style="position:absolute;top:0;left:-4px;bottom:0;width:4px;background:linear-gradient(180deg,var(--warning),#fbbf24);border-radius:var(--radius-lg) 0 0 var(--radius-lg)"></div>
        <div class="item-head">
            <h2>{{gc.group}}</h2>
            <span class="tag" style="background:var(--warning-bg);color:#92400e">{{gc.count}} items</span>
            <span class="tag">{{selected_metric}}</span>
            <span class="tag" style="background:#f5f3ff;color:#7c3aed">Group Total</span>
        </div>
        <div id="chart_grp_{{loop.index0}}" style="width:100%;height:440px"></div>
        <script>Plotly.newPlot('chart_grp_{{loop.index0}}',{{gc.chart_data|safe}}.data,{{gc.chart_data|safe}}.layout,{responsive:true})</script>
    </div>
    {% endfor %}
    {% endif %}

    {% for item_data in results %}
    <div class="card item-card fade-up">
        <div class="item-head">
            <h2>{{item_data.item}}</h2>
            <span class="tag">{{item_data.group}}</span>
            <span class="tag">{{selected_metric}}</span>
            {% if item_data.best_method %}
            <span class="best-badge"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5"><path d="M22 11.08V12a10 10 0 11-5.93-9.14"/><polyline points="22 4 12 14.01 9 11.01"/></svg> Best: {{item_data.best_method}}</span>
            {% endif %}
        </div>
        <div id="chart_{{loop.index0}}" style="width:100%;height:440px"></div>
        <script>Plotly.newPlot('chart_{{loop.index0}}',{{item_data.chart_data|safe}}.data,{{item_data.chart_data|safe}}.layout,{responsive:true})</script>

        <div class="section-lbl">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="7" height="7"/><rect x="14" y="3" width="7" height="7"/><rect x="3" y="14" width="7" height="7"/><rect x="14" y="14" width="7" height="7"/></svg>
            Forecast Values
        </div>
        <div class="scroll-tbl">
        <table>
            <thead><tr><th>Month</th>{% for m in selected_methods %}<th>{{m}}</th>{% endfor %}</tr></thead>
            <tbody>
            {% for i in range(forecast_months|length) %}
            <tr><td><strong>{{forecast_months[i]}}</strong></td>
            {% for m in selected_methods %}
                <td class="fc-val">{{item_data.forecasts.get(m, [0]*9)[i]|round(2)|string|replace('.0','') if item_data.forecasts.get(m) else 'N/A'}}</td>
            {% endfor %}</tr>
            {% endfor %}
            </tbody>
        </table>
        </div>

        {% if item_data.accuracy %}
        <div class="section-lbl">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M22 11.08V12a10 10 0 11-5.93-9.14"/><polyline points="22 4 12 14.01 9 11.01"/></svg>
            Backtest Accuracy
            <span style="font-weight:400;font-size:.7rem;color:var(--text-4);margin-left:4px">(last 3 months held out)</span>
        </div>
        <div class="scroll-tbl">
        <table>
            <thead><tr><th>Method</th><th>MAE</th><th>MAPE (%)</th><th>Rating</th></tr></thead>
            <tbody>
            {% for acc in item_data.accuracy %}
            <tr class="{% if acc.method == item_data.best_method %}best-row{% endif %}">
                <td style="font-weight:500">{{acc.method}}{% if acc.method == item_data.best_method %} <span class="best-badge" style="font-size:.6rem;padding:2px 7px">BEST</span>{% endif %}</td>
                <td>{{acc.mae if acc.mae is not none else 'N/A'}}</td>
                <td>{{acc.mape if acc.mape is not none else 'N/A'}}</td>
                <td>{% if acc.rating == 'Good' %}<span class="badge badge-good">Good</span>
                    {% elif acc.rating == 'Fair' %}<span class="badge badge-fair">Fair</span>
                    {% elif acc.rating == 'Poor' %}<span class="badge badge-poor">Poor</span>
                    {% else %}-{% endif %}</td>
            </tr>
            {% endfor %}
            </tbody>
        </table>
        </div>
        {% endif %}
    </div>
    {% endfor %}

    {% if not results %}
    <div class="card fade-up"><div class="empty"><p>Select items and methods, then click <strong>Apply Filters</strong>.</p></div></div>
    {% endif %}

    <!-- METHOD EXPLANATIONS -->
    <div class="card fade-up">
        <h2 style="font-size:1.05rem;font-weight:700;color:var(--text);margin-bottom:16px;display:flex;align-items:center;gap:8px">
            <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="var(--primary)" stroke-width="2"><circle cx="12" cy="12" r="10"/><path d="M9.09 9a3 3 0 015.83 1c0 2-3 3-3 3"/><line x1="12" y1="17" x2="12.01" y2="17"/></svg>
            Forecasting Methods
        </h2>
        <div class="tabs" id="methodTabs">
        {% for mname in method_descriptions %}
            <div class="tab {% if loop.first %}active{% endif %}" onclick="showTab(this,'mt_{{loop.index0}}')">{{mname}}</div>
        {% endfor %}
        </div>
        {% for mname, mdesc in method_descriptions.items() %}
        <div class="tab-content {% if loop.first %}active{% endif %}" id="mt_{{loop.index0}}">
            <div class="method-desc">{{mdesc|safe}}</div>
        </div>
        {% endfor %}

        <div class="guide">
            <h3>Which Method to Use</h3>
            <div class="scroll-tbl">
            <table>
                <thead><tr><th>Sales Pattern</th><th>Recommended Method</th></tr></thead>
                <tbody>
                <tr><td>Strong seasonal peaks</td><td><strong>Holt-Winters (Multiplicative)</strong> or Ensemble</td></tr>
                <tr><td>Steady growth, no seasonality</td><td><strong>Exponential Smoothing</strong> or Moving Average</td></tr>
                <tr><td>Stable with mild seasonality</td><td><strong>Linear Trend + Seasonality</strong></td></tr>
                <tr><td>New product (few months)</td><td><strong>Moving Average (3M)</strong></td></tr>
                <tr><td>Unsure / mixed</td><td><strong>Ensemble (Best-of-3)</strong> - safest default</td></tr>
                </tbody>
            </table>
            </div>
            <div class="guide-note">
                <strong>MAE</strong> = avg absolute difference (lower is better, same units as data).<br>
                <strong>MAPE</strong> = avg % error. Below 30% = Good, 30-60% = Fair, above 60% = Poor.<br>
                <strong>Best Method</strong> = automatically selected per item (lowest MAPE). Highlighted green in accuracy table and included in Excel "Best Method" tab.
            </div>
        </div>
    </div>
</div>
</div>
</form>
</div>

{% endif %}

<footer>
    Farmley Forecast Dashboard v6.0 &bull; Flask + Plotly &bull; <a href="https://github.com/Rameshwarnaik013/farmley-forecast-dashboard" target="_blank">GitHub</a>
</footer>

<script>
function filterCB(input, listId){
    var q=input.value.toLowerCase();
    var labels=document.getElementById(listId).querySelectorAll('label');
    labels.forEach(function(lb){
        var t=lb.querySelector('.cname').textContent.toLowerCase();
        lb.classList.toggle('hidden', t.indexOf(q)===-1);
    });
}
function checkAll(listId){
    document.getElementById(listId).querySelectorAll('label:not(.hidden) input[type=checkbox]').forEach(function(cb){cb.checked=true});
    updateInfo(listId);
}
function uncheckAll(listId){
    document.getElementById(listId).querySelectorAll('input[type=checkbox]').forEach(function(cb){cb.checked=false});
    updateInfo(listId);
}
function updateInfo(listId){
    var total=document.getElementById(listId).querySelectorAll('input[type=checkbox]').length;
    var checked=document.getElementById(listId).querySelectorAll('input[type=checkbox]:checked').length;
    var el=document.getElementById(listId+'_info');
    if(el) el.textContent=checked+' of '+total+' selected';
}
function showTab(el,id){
    el.parentElement.querySelectorAll('.tab').forEach(function(t){t.classList.remove('active')});
    el.classList.add('active');
    var card=el.closest('.card');
    card.querySelectorAll('.tab-content').forEach(function(c){c.classList.remove('active')});
    card.querySelector('#'+id).classList.add('active');
}
['grp','itm','mth'].forEach(function(id){
    var el=document.getElementById(id);
    if(el){
        updateInfo(id);
        el.addEventListener('change', function(){updateInfo(id)});
    }
});
</script>
</body>
</html>"""


def md_to_html(text):
    text = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', text)
    text = re.sub(r'`(.+?)`', r'<code>\1</code>', text)
    text = text.replace('\n\n', '<br><br>').replace('\n-', '<br>*')
    return text


@app.route("/")
def index():
    if "df_json" not in _cached_data:
        return render_template_string(TEMPLATE, data_loaded=False)
    return redirect("/forecast")


@app.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f:
        return redirect("/")
    df = pd.read_excel(f)
    meta = {"Item_Name", "New MIS ITEM Group", "Metric"}
    month_cols = [c for c in df.columns if c not in meta]
    df[month_cols] = df[month_cols].fillna(0)
    for c in month_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    _cached_data["df_json"] = df.to_json()
    _cached_data["month_cols"] = month_cols
    return redirect("/forecast")


@app.route("/forecast", methods=["GET", "POST"])
def forecast():
    if "df_json" not in _cached_data:
        return redirect("/")

    df = pd.read_json(io.StringIO(_cached_data["df_json"]))
    month_cols = _cached_data["month_cols"]
    forecast_months = compute_forecast_months(month_cols)

    all_groups = sorted([g for g in df["New MIS ITEM Group"].dropna().unique()])
    all_items = sorted(df["Item_Name"].dropna().unique().tolist())
    all_metrics = sorted(df["Metric"].dropna().unique().tolist())
    all_methods = list(FORECAST_METHODS.keys())

    if request.method == "POST":
        selected_groups = request.form.getlist("groups")
        selected_items = request.form.getlist("items")
        selected_metric = request.form.get("metric", all_metrics[0] if all_metrics else "")
        selected_methods = request.form.getlist("methods")
    else:
        selected_groups = all_groups
        selected_items = all_items
        selected_metric = all_metrics[0] if all_metrics else ""
        selected_methods = list(FORECAST_METHODS.keys())

    if not selected_groups:
        selected_groups = all_groups
    if not selected_methods:
        selected_methods = list(FORECAST_METHODS.keys())

    group_items = df[df["New MIS ITEM Group"].isin(selected_groups)]["Item_Name"].unique().tolist()
    filtered_items = sorted(set(all_items) & set(group_items))
    selected_items = [it for it in selected_items if it in filtered_items]
    if not selected_items:
        selected_items = filtered_items[:5]

    results = []
    for item_name in selected_items[:20]:
        row = df[(df["Item_Name"] == item_name) & (df["Metric"] == selected_metric)]
        if row.empty:
            continue
        hist = row.iloc[0][month_cols].values.astype(float)
        series = pd.Series(hist, index=pd.date_range("2020-01-01", periods=len(month_cols), freq="MS"))
        group = row.iloc[0].get("New MIS ITEM Group", "—")

        forecasts = {}
        accuracy = []
        for m in selected_methods:
            fn = FORECAST_METHODS[m]
            try:
                fc = [float(v) for v in fn(series)[:9]]
                forecasts[m] = fc
            except Exception:
                forecasts[m] = [0.0] * 9
            mae, mape = backtest_method(hist, fn)
            rating = "—"
            if mape is not None:
                rating = "Good" if mape < 30 else ("Fair" if mape < 60 else "Poor")
            accuracy.append({"method": m, "mae": mae, "mape": mape, "rating": rating})

        # Find best method (lowest MAPE) for this item
        best_method_name = None
        best_mape_val = float("inf")
        for acc_entry in accuracy:
            if acc_entry["mape"] is not None and acc_entry["mape"] < best_mape_val:
                best_mape_val = acc_entry["mape"]
                best_method_name = acc_entry["method"]

        chart_json = build_chart_json(item_name, selected_metric, hist, forecasts, month_cols, forecast_months)
        results.append({
            "item": item_name, "group": group,
            "chart_data": chart_json,
            "forecasts": forecasts,
            "accuracy": accuracy,
            "best_method": best_method_name,
        })

    # Build overall aggregate chart (sum across all selected items)
    overall_chart = None
    if results:
        agg_hist = np.zeros(len(month_cols))
        agg_forecasts = {m: np.zeros(9) for m in selected_methods}
        for rd in results:
            row = df[(df["Item_Name"] == rd["item"]) & (df["Metric"] == selected_metric)]
            if not row.empty:
                agg_hist += row.iloc[0][month_cols].values.astype(float)
            for m in selected_methods:
                agg_forecasts[m] += np.array(rd["forecasts"].get(m, [0.0]*9))
        agg_fc_dict = {m: agg_forecasts[m].tolist() for m in selected_methods}
        overall_chart = build_chart_json(
            f"All Selected Items ({len(results)})", selected_metric,
            agg_hist, agg_fc_dict, month_cols, forecast_months
        )

    # Build group-level aggregate charts
    group_charts = []
    if results:
        groups_in_results = {}
        for rd in results:
            g = rd["group"]
            if g not in groups_in_results:
                groups_in_results[g] = {"hist": np.zeros(len(month_cols)),
                                        "fc": {m: np.zeros(9) for m in selected_methods},
                                        "count": 0}
            row = df[(df["Item_Name"] == rd["item"]) & (df["Metric"] == selected_metric)]
            if not row.empty:
                groups_in_results[g]["hist"] += row.iloc[0][month_cols].values.astype(float)
            for m in selected_methods:
                groups_in_results[g]["fc"][m] += np.array(rd["forecasts"].get(m, [0.0]*9))
            groups_in_results[g]["count"] += 1
        for gname, gdata in sorted(groups_in_results.items()):
            if gdata["count"] < 2:
                continue  # skip groups with only 1 item (already shown in item view)
            gc_dict = {m: gdata["fc"][m].tolist() for m in selected_methods}
            gc_json = build_chart_json(
                f"{gname} ({gdata['count']} items)", selected_metric,
                gdata["hist"], gc_dict, month_cols, forecast_months
            )
            group_charts.append({"group": gname, "count": gdata["count"], "chart_data": gc_json})

    desc_html = {k: md_to_html(v) for k, v in METHOD_DESCRIPTIONS.items()}

    return render_template_string(TEMPLATE,
        data_loaded=True,
        all_groups=all_groups, selected_groups=selected_groups,
        all_items=filtered_items, selected_items=selected_items,
        all_metrics=all_metrics, selected_metric=selected_metric,
        all_methods=all_methods, selected_methods=selected_methods,
        forecast_months=forecast_months, results=results,
        overall_chart=overall_chart, group_charts=group_charts,
        total_items=len(all_items), last_month=month_cols[-1],
        n_months=len(month_cols),
        method_descriptions=desc_html,
    )


@app.route("/download")
def download():
    if "df_json" not in _cached_data:
        return redirect("/")
    df = pd.read_json(io.StringIO(_cached_data["df_json"]))
    month_cols = _cached_data["month_cols"]
    forecast_months = compute_forecast_months(month_cols)
    methods_str = request.args.get("methods", "Ensemble (Best-of-3)")
    selected_methods = [m.strip() for m in methods_str.split(",") if m.strip() in FORECAST_METHODS]
    if not selected_methods:
        selected_methods = ["Ensemble (Best-of-3)"]
    buf = create_excel_output(df, month_cols, forecast_months, selected_methods)
    return send_file(buf, download_name="Farmley_Forecast_Output.xlsx",
                     as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(debug=True, port=5000)
