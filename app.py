import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from forecasting import (
    moving_average_forecast,
    weighted_moving_average_forecast,
    exponential_smoothing_forecast,
    holt_winters_forecast,
    linear_trend_forecast,
    seasonal_naive_forecast,
    seasonal_decomposition_forecast,
    ensemble_forecast,
    METHOD_DESCRIPTIONS,
)

st.set_page_config(
    page_title="Farmley Sales Forecast Dashboard",
    page_icon="🔮",
    layout="wide",
)

FORECAST_METHODS = {
    "Seasonal Naive": seasonal_naive_forecast,
    "Moving Average (3M)": lambda s: moving_average_forecast(s, window=3),
    "Moving Average (6M)": lambda s: moving_average_forecast(s, window=6),
    "Weighted Moving Average": weighted_moving_average_forecast,
    "Exponential Smoothing": exponential_smoothing_forecast,
    "Holt-Winters (Additive)": lambda s: holt_winters_forecast(s, seasonal="add"),
    "Holt-Winters (Multiplicative)": lambda s: holt_winters_forecast(s, seasonal="mul"),
    "Linear Trend + Seasonality": linear_trend_forecast,
    "Seasonal Decomposition": seasonal_decomposition_forecast,
    "Ensemble (Best-of-3)": ensemble_forecast,
}

MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def parse_month_col(col_name):
    """Parse 'Apr-2025' or 'Apr-26' into a (year, month_index) tuple."""
    m = re.match(r"([A-Za-z]{3})-(\d{2,4})", str(col_name))
    if not m:
        return None
    abbr, yr = m.group(1), int(m.group(2))
    if yr < 100:
        yr += 2000
    try:
        mi = MONTH_ABBR.index(abbr.capitalize())
    except ValueError:
        return None
    return (yr, mi)


def next_month(year, mi):
    """Given (year, month_index 0-11), return next month's (year, mi)."""
    mi += 1
    if mi > 11:
        mi = 0
        year += 1
    return year, mi


def month_label(year, mi):
    """Format (year, month_index) as 'Apr-2026'."""
    return f"{MONTH_ABBR[mi]}-{year}"


def compute_forecast_months(month_cols, n_forecast=9):
    """Determine the next n_forecast month labels after the last data column."""
    last_col = month_cols[-1]
    parsed = parse_month_col(last_col)
    if parsed is None:
        yr, mi = 2026, 3
    else:
        yr, mi = parsed
    result = []
    for _ in range(n_forecast):
        yr, mi = next_month(yr, mi)
        result.append(month_label(yr, mi))
    return result


@st.cache_data
def load_data(file):
    df = pd.read_excel(file)
    meta_cols = {"Item_Name", "New MIS ITEM Group", "Metric"}
    month_cols = [c for c in df.columns if c not in meta_cols]
    df[month_cols] = df[month_cols].fillna(0)
    for c in month_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return df, month_cols


def run_forecast(series, method_name):
    return FORECAST_METHODS[method_name](series)


def backtest_method(hist_values, method_name, n_test=3):
    """Train on all-but-last-n_test months, predict n_test, return MAE and MAPE."""
    if len(hist_values) < n_test + 3:
        return None, None
    train = hist_values[:-n_test]
    test = hist_values[-n_test:]
    train_series = pd.Series(train, index=pd.date_range("2020-01-01", periods=len(train), freq="MS"))
    try:
        pred = FORECAST_METHODS[method_name](train_series)[:n_test]
        pred = np.array([float(v) for v in pred])
        mae = float(np.mean(np.abs(pred - test)))
        safe_test = np.where(test == 0, 1, test)
        mape = float(np.mean(np.abs((pred - test) / safe_test))) * 100
        return round(mae, 2), round(mape, 1)
    except Exception:
        return None, None


def build_chart(item_name, metric, hist_values, forecasts_dict, month_cols, forecast_months):
    fig = go.Figure()
    x_actual = list(month_cols)
    fig.add_trace(go.Scatter(
        x=x_actual, y=[float(v) for v in hist_values],
        mode="lines+markers", name="Actual",
        line=dict(color="#1f77b4", width=3),
        marker=dict(size=8),
    ))
    colors = [
        "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
        "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf", "#ff6600",
    ]
    all_x = x_actual + forecast_months
    for i, (method, vals) in enumerate(forecasts_dict.items()):
        all_y = [float(v) for v in hist_values] + [float(v) for v in vals]
        fig.add_trace(go.Scatter(
            x=all_x, y=all_y,
            mode="lines+markers", name=method,
            line=dict(color=colors[i % len(colors)], width=2, dash="dash"),
            marker=dict(size=6),
        ))
    last_idx = len(x_actual) - 1
    fig.add_shape(
        type="line", x0=last_idx, x1=last_idx,
        y0=0, y1=1, yref="paper",
        line=dict(color="gray", width=2, dash="dot"),
    )
    fig.add_annotation(
        x=last_idx, y=1.05, yref="paper",
        text="← Actual | Forecast →",
        showarrow=False, font=dict(size=11, color="gray"),
    )
    fig.update_layout(
        title=f"{item_name} — {metric}",
        xaxis_title="Month", yaxis_title=metric,
        hovermode="x unified", height=520,
        legend=dict(orientation="h", yanchor="bottom", y=-0.35),
        margin=dict(b=120),
    )
    return fig


def create_excel_output(df, month_cols, forecast_months, forecasts_all, selected_methods):
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    forecast_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    actual_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for method in selected_methods:
        ws = wb.create_sheet(title=method[:31])
        headers = ["Item_Name", "Group", "Metric"] + list(month_cols) + forecast_months + ["MAE", "MAPE (%)"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            item = row["Item_Name"]
            group = row.get("New MIS ITEM Group", "")
            metric = row["Metric"]
            ws.cell(row=r_idx, column=1, value=item).border = thin_border
            ws.cell(row=r_idx, column=2, value=group).border = thin_border
            ws.cell(row=r_idx, column=3, value=metric).border = thin_border

            hist_vals = []
            for c_i, m in enumerate(month_cols, 4):
                val = float(row[m])
                cell = ws.cell(row=r_idx, column=c_i, value=val)
                cell.fill = actual_fill
                cell.border = thin_border
                cell.number_format = '#,##0.00'
                hist_vals.append(val)

            key = (item, metric, method)
            fvals = forecasts_all.get(key, {}).get("forecast", [0.0] * 9)
            mae_val = forecasts_all.get(key, {}).get("mae")
            mape_val = forecasts_all.get(key, {}).get("mape")

            fc_start_col = 4 + len(month_cols)
            for fc_i, fv in enumerate(fvals):
                cell = ws.cell(row=r_idx, column=fc_start_col + fc_i, value=round(float(fv), 2))
                cell.fill = forecast_fill
                cell.border = thin_border
                cell.number_format = '#,##0.00'
                cell.comment = openpyxl.comments.Comment(
                    _get_formula_comment(method, hist_vals, fc_i), "Forecast Engine"
                )

            mae_col = fc_start_col + len(forecast_months)
            mape_col = mae_col + 1
            if mae_val is not None:
                ws.cell(row=r_idx, column=mae_col, value=mae_val).border = thin_border
                ws.cell(row=r_idx, column=mae_col).number_format = '#,##0.00'
            if mape_val is not None:
                ws.cell(row=r_idx, column=mape_col, value=mape_val).border = thin_border
                ws.cell(row=r_idx, column=mape_col).number_format = '#,##0.0'

        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 18
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    formula_ws = wb.create_sheet(title="Method Explanations")
    formula_ws.column_dimensions["A"].width = 30
    formula_ws.column_dimensions["B"].width = 100
    formula_ws.cell(row=1, column=1, value="Method").font = Font(bold=True, size=12)
    formula_ws.cell(row=1, column=2, value="Explanation & Formula").font = Font(bold=True, size=12)
    for i, (mname, mdesc) in enumerate(METHOD_DESCRIPTIONS.items(), 2):
        formula_ws.cell(row=i, column=1, value=mname).font = Font(bold=True)
        formula_ws.cell(row=i, column=2, value=mdesc)
        formula_ws.row_dimensions[i].height = 60

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _get_formula_comment(method, hist_vals, fc_index):
    if "Moving Average (3M)" in method:
        if fc_index == 0:
            src = hist_vals[-3:]
            return f"MA(3) = AVG of last 3 actuals = ({' + '.join(f'{v:.1f}' for v in src)}) / 3 = {np.mean(src):.1f}"
        return "MA(3) = rolling average of previous 3 values (including prior forecasts)"
    elif "Moving Average (6M)" in method:
        if fc_index == 0:
            src = hist_vals[-6:]
            return f"MA(6) = AVG of last 6 actuals = ({' + '.join(f'{v:.1f}' for v in src)}) / 6 = {np.mean(src):.1f}"
        return "MA(6) = rolling average of previous 6 values"
    elif "Weighted" in method:
        return "WMA: weights 0.30, 0.25, 0.20, 0.15, 0.10 for last 5 months"
    elif "Exponential" in method:
        return "SES: F(t) = alpha*A(t-1) + (1-alpha)*F(t-1), alpha optimized via SSE"
    elif "Holt-Winters" in method:
        return "Holt-Winters: Level + Trend + Seasonal. Period=12. Params via MLE."
    elif "Linear Trend" in method:
        return "Linear Trend: Y = (a + b*t) * S(month). OLS trend, seasonal index = avg(actual/trend)."
    elif "Seasonal Naive" in method:
        if fc_index < len(hist_vals):
            return f"Seasonal Naive: Same month last year = {hist_vals[fc_index]:.1f}"
        return "Seasonal Naive: Repeats same month from previous year"
    elif "Decomposition" in method:
        return "STL: Trend (LOESS) + Seasonal extracted, trend extrapolated"
    elif "Ensemble" in method:
        return "Ensemble: Median of Seasonal Naive, HW-Mul, Linear Trend"
    return f"{method} forecast"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN APP
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

st.title("Farmley Sales Forecast Dashboard")
st.markdown("Upload your item-wise sales order Excel to generate next-9-month forecasts using multiple methods.")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx", "xls"])
if uploaded_file is None:
    import os
    default_path = os.environ.get("FORECAST_DEFAULT_FILE", r"C:\Users\Admin\Downloads\Forecast_FARMLEY.xlsx")
    if os.path.exists(default_path):
        df, month_cols = load_data(default_path)
        st.info(f"Using default data file. Upload a new file above to replace it.")
    else:
        st.warning("Please upload your Forecast Excel file to get started.")
        st.stop()
else:
    df, month_cols = load_data(uploaded_file)
forecast_months = compute_forecast_months(month_cols, n_forecast=9)
last_data_month = month_cols[-1]

st.success(f"Loaded {len(df)} rows — {len(df['Item_Name'].unique())} items, data through **{last_data_month}**. Forecasting: **{forecast_months[0]}** to **{forecast_months[-1]}**.")

# ─── SIDEBAR FILTERS ────────────────────────────────────────────────────────────
st.sidebar.header("Filters")

# --- Group multiselect with select/deselect all ---
all_groups = sorted([g for g in df["New MIS ITEM Group"].dropna().unique()])

st.sidebar.markdown("**Product Group**")
gcol1, gcol2 = st.sidebar.columns(2)
if gcol1.button("Select All Groups", use_container_width=True):
    st.session_state["sel_groups"] = all_groups
if gcol2.button("Clear Groups", use_container_width=True):
    st.session_state["sel_groups"] = []

selected_groups = st.sidebar.multiselect(
    "Search & select groups",
    options=all_groups,
    default=st.session_state.get("sel_groups", all_groups),
    key="group_multiselect",
    label_visibility="collapsed",
)
st.session_state["sel_groups"] = selected_groups

if not selected_groups:
    filtered_df = df
    st.sidebar.caption("No groups selected — showing all items")
else:
    filtered_df = df[df["New MIS ITEM Group"].isin(selected_groups)]

# --- Item multiselect with select/deselect all ---
all_items = sorted(filtered_df["Item_Name"].dropna().unique())

st.sidebar.markdown("**Item Name**")
icol1, icol2 = st.sidebar.columns(2)
if icol1.button("Select All Items", use_container_width=True):
    st.session_state["sel_items"] = all_items
if icol2.button("Clear Items", use_container_width=True):
    st.session_state["sel_items"] = []

prev_items = st.session_state.get("sel_items", all_items[:1])
valid_prev = [i for i in prev_items if i in all_items]
if not valid_prev:
    valid_prev = all_items[:1]

selected_items = st.sidebar.multiselect(
    "Search & select items",
    options=all_items,
    default=valid_prev,
    key="item_multiselect",
    label_visibility="collapsed",
)
st.session_state["sel_items"] = selected_items

if not selected_items:
    st.warning("Select at least one item from the sidebar.")
    st.stop()

# --- Metric selector ---
metrics = sorted(df[df["Item_Name"].isin(selected_items)]["Metric"].dropna().unique())
selected_metric = st.sidebar.selectbox("Metric", metrics)

# --- Forecast methods with select/deselect all ---
st.sidebar.divider()
st.sidebar.header("Forecast Methods")
all_method_names = list(FORECAST_METHODS.keys())

mcol1, mcol2 = st.sidebar.columns(2)
if mcol1.button("Select All Methods", use_container_width=True):
    st.session_state["sel_methods"] = all_method_names
if mcol2.button("Clear Methods", use_container_width=True):
    st.session_state["sel_methods"] = []

default_methods = st.session_state.get("sel_methods", [
    "Seasonal Naive", "Holt-Winters (Multiplicative)",
    "Linear Trend + Seasonality", "Ensemble (Best-of-3)",
])
valid_defaults = [m for m in default_methods if m in all_method_names]

selected_methods = st.sidebar.multiselect(
    "Search & select methods",
    options=all_method_names,
    default=valid_defaults,
    key="method_multiselect",
    label_visibility="collapsed",
)
st.session_state["sel_methods"] = selected_methods

if not selected_methods:
    st.warning("Select at least one forecast method from the sidebar.")
    st.stop()

# ─── PER-ITEM FORECAST + DISPLAY ───────────────────────────────────────────────
for selected_item in selected_items:
    item_rows = df[(df["Item_Name"] == selected_item) & (df["Metric"] == selected_metric)]
    if item_rows.empty:
        st.warning(f"No {selected_metric} data for **{selected_item}**.")
        continue

    hist_values = item_rows.iloc[0][month_cols].values.astype(float)
    series = pd.Series(hist_values, index=pd.date_range("2020-01-01", periods=len(month_cols), freq="MS"))

    forecasts = {}
    accuracy = {}
    for method in selected_methods:
        try:
            fc = run_forecast(series, method)
            forecasts[method] = [float(v) for v in fc[:9]]
        except Exception as e:
            st.warning(f"{method} failed for {selected_item}: {e}")
            continue
        mae, mape = backtest_method(hist_values, method)
        accuracy[method] = {"MAE": mae, "MAPE (%)": mape}

    st.subheader(selected_item)
    group_name = item_rows.iloc[0].get("New MIS ITEM Group", "—")
    c1, c2, c3 = st.columns(3)
    c1.metric("Group", group_name)
    c2.metric("Metric", selected_metric)
    c3.metric("Data Months", len(month_cols))

    if forecasts:
        fig = build_chart(selected_item, selected_metric, hist_values, forecasts, list(month_cols), forecast_months)
        st.plotly_chart(fig, use_container_width=True)

    # Forecast + accuracy table
    table_data = {"Month": forecast_months}
    for method, vals in forecasts.items():
        table_data[method] = [round(v, 2) for v in vals]
    fc_df = pd.DataFrame(table_data)
    st.dataframe(fc_df, use_container_width=True, hide_index=True)

    # Accuracy table
    if accuracy:
        acc_rows = []
        for method, acc in accuracy.items():
            mape_val = acc.get("MAPE (%)")
            rating = "—"
            if mape_val is not None:
                rating = "Good" if mape_val < 30 else ("Fair" if mape_val < 60 else "Poor")
            acc_rows.append({
                "Method": method,
                "MAE": acc.get("MAE", "N/A"),
                "MAPE (%)": mape_val if mape_val is not None else "N/A",
                "Rating": rating,
            })
        acc_df = pd.DataFrame(acc_rows)
        st.markdown("**Backtest Accuracy** _(trained on all-but-last-3 months, tested on last 3)_")
        st.dataframe(acc_df, use_container_width=True, hide_index=True)

    st.divider()

# ─── BULK DOWNLOAD ──────────────────────────────────────────────────────────────
st.subheader("Download Full Forecast Excel")
st.markdown(f"Forecasts for **all {len(df['Item_Name'].unique())} items** with selected methods. Each method gets its own sheet with MAE/MAPE columns and formula comments.")

if st.button("Generate Full Forecast Excel", type="primary"):
    progress = st.progress(0, text="Computing forecasts...")
    forecasts_all = {}
    total = len(df)
    for idx, (_, row) in enumerate(df.iterrows()):
        item = row["Item_Name"]
        metric = row["Metric"]
        vals = row[month_cols].values.astype(float)
        s = pd.Series(vals, index=pd.date_range("2020-01-01", periods=len(month_cols), freq="MS"))
        for method in selected_methods:
            try:
                fc = run_forecast(s, method)
                fc_vals = [float(v) for v in fc[:9]]
            except Exception:
                fc_vals = [0.0] * 9
            mae, mape = backtest_method(vals, method)
            forecasts_all[(item, metric, method)] = {
                "forecast": fc_vals, "mae": mae, "mape": mape,
            }
        if idx % 50 == 0:
            progress.progress(min(idx / total, 1.0), text=f"Processing {idx}/{total} rows...")

    progress.progress(1.0, text="Building Excel...")
    buf = create_excel_output(df, month_cols, forecast_months, forecasts_all, selected_methods)
    progress.empty()
    st.download_button(
        "Download Forecast Excel",
        data=buf,
        file_name="Farmley_Forecast_Output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.openxmlformats",
    )
    st.success("Forecast generated for all items!")

# ─── GROUP SUMMARY ──────────────────────────────────────────────────────────────
st.subheader("Group-Level Summary")
if selected_groups:
    group_df = df[df["New MIS ITEM Group"].isin(selected_groups)]
    display_group = ", ".join(selected_groups[:3]) + ("..." if len(selected_groups) > 3 else "")
else:
    group_df = df
    display_group = "All Groups"

group_metric_df = group_df[group_df["Metric"] == selected_metric]
group_totals = group_metric_df[month_cols].sum()

fig_group = go.Figure()
fig_group.add_trace(go.Bar(
    x=list(month_cols), y=group_totals.values.tolist(),
    name="Actual", marker_color="#1f77b4",
))
fig_group.update_layout(
    title=f"Group Total: {display_group} — {selected_metric}",
    xaxis_title="Month", yaxis_title=selected_metric,
    height=400,
)
st.plotly_chart(fig_group, use_container_width=True)

# ─── METHOD EXPLANATIONS ────────────────────────────────────────────────────────
st.divider()
st.header("Forecasting Methods — Logic & Approach")
st.markdown("Understanding each method helps you pick the right one for each product's sales pattern.")

tab_names = list(METHOD_DESCRIPTIONS.keys())
tabs = st.tabs(tab_names)
for tab, method_name in zip(tabs, tab_names):
    with tab:
        st.markdown(METHOD_DESCRIPTIONS[method_name])

st.divider()
st.markdown("""
**Quick Guide — Which Method to Use:**

| Sales Pattern | Recommended Method |
|---|---|
| Strong seasonal peaks (festive months) | Holt-Winters (Multiplicative) or Ensemble |
| Steady growth, no seasonality | Exponential Smoothing or Moving Average |
| Stable with mild seasonality | Linear Trend + Seasonality |
| New product (few months of data) | Moving Average (3M) |
| Unsure / mixed pattern | **Ensemble (Best-of-3)** — safest default |

The **Backtest Accuracy** table above each item shows which method fits best (lowest MAPE wins).

**Accuracy Metrics Explained:**
- **MAE** (Mean Absolute Error): Average absolute difference between predicted and actual values. Lower = better. In the same units as your data.
- **MAPE** (Mean Absolute Percentage Error): Average percentage error. Lower = better. <30% is Good, 30-60% is Fair, >60% is Poor.
""")

st.divider()
st.caption("Farmley Sales Forecast Dashboard v2.0 | Built with Streamlit + Plotly")
